import csv
import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def getattr_recursive(obj, attr):
    """
    Support accessing 'user.username' string format.
    """
    try:
        value = obj
        for part in attr.split('.'):
            if hasattr(value, part):
                value = getattr(value, part)
                if callable(value):
                    value = value()
            else:
                return None
        return value
    except Exception:
        return None

def export_data(queryset, format_type, resource_name, columns, headers):
    """
    columns: list of model attributes/methods to fetch (e.g. ['id', 'amount', 'user.username'])
    headers: list of strings for header row
    """
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"{resource_name}_{timestamp}.{format_type}"

    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add BOM for Excel compatibility with UTF-8
        response.write(u'\ufeff'.encode('utf8'))
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for obj in queryset:
            row = []
            for col in columns:
                val = getattr_recursive(obj, col)
                row.append(str(val) if val is not None else '')
            writer.writerow(row)
        return response

    elif format_type == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = resource_name[:30]
        
        # Header with Bold Font
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        # Data
        for obj in queryset:
            row = []
            for col in columns:
                val = getattr_recursive(obj, col)
                row.append(str(val) if val is not None else '')
            ws.append(row)
            
        # Auto-adjust column width
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                try:
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        wb.save(response)
        return response

    elif format_type == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        template_path = 'exports/pdf_template.html'
        
        # Prepare context data
        data_list = []
        for obj in queryset:
            row = []
            for col in columns:
                val = getattr_recursive(obj, col)
                row.append(val if val is not None else '')
            data_list.append(row)
            
        context = {
            'resource_name': resource_name,
            'date': datetime.datetime.now(),
            'headers': headers,
            'data_list': data_list,
        }
        
        template = get_template(template_path)
        html = template.render(context)
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse(f'We had some errors <pre>{html}</pre>')
        return response

    return HttpResponse("Invalid Format", status=400)
